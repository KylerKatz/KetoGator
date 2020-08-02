import xlrd
from openpyxl import *
import os
import pandas

# Creates appropriate directories for new patient or existing patient who doesn't have them
def setNewPatient(MrNum):
    path = r"Current Patients\\" + MrNum + r"\\DataBases\\Data\\"

    # If path doesn't exist, the function executes, otherwise nothing happens
    if not os.path.isdir(path):

        # Used to make appropriate directories
        switch = {
            0: MrNum,
            1: "\DataBases",
            2: "\Data",
        }
        dir = "Current Patients\\"

        # Create necessary directories
        for i in range(3):
            dir = dir + switch.get(i)
            os.mkdir(dir)

# Returns list of all patients
def getAllPatients():
    cwd = os.getcwd() + "\\Current Patients"
    return (os.listdir(cwd))

# Returns list of all excel sheets a patient has
def getPatientExcels(patient):
    return (os.listdir("Current Patients\\" + patient + "\\DataBases\\Data\\"))

# Returns list of all graphs a patient has
def getPatientGraphs(patient):
    graphs = []

    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Anthropometrics_Source.xlsx"

    if os.path.isfile(path):
        graphs.append("Anthropometrics")
        
    path2 = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Med_Data_Source.xlsx"

    if os.path.isfile(path) and os.path.isfile(path2):
        graphs.append("Med Load")


    return (graphs)

################### Save Methods for excel sheets ################### 

def saveAlertness(patient, MRNumber, Date, DayType, Alertness, Activity, Development, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Alertness_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb['Sheet1']
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Create appropriate columns
        ws.append(
            ["MRNumber", "Date", "Day_Type", "Alertness", "Activity", "Development", "Entered", "Audited", "Comments"])

    # Add data
    ws.append([MRNumber, Date, DayType, Alertness, Activity, Development, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveAnthropometrics(patient, MRNumber, Date, DayType, Source, CP, PA, Ht, Wt, HC, UAC, TSF, SSF, USF, SIS, MBSF, UC,
                        Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Anthropometrics_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb['Anthropometrics']
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Anthropometrics"

        # Create appropriate columns
        ws.append(
            ["MRNumber", "Date", "Day_Type", "Source", "CP", "PA", "Ht", "Wt", "HC", "UAC", "TSF", "SSF", "USF", "SIS",
             "MBSF", "UC", "R", "X", "Entered", "Audited", "Comments"])

    # Add data
    ws.append(
        [MRNumber, Date, DayType, Source, CP, PA, Ht, Wt, HC, UAC, TSF, SSF, USF, SIS, MBSF, UC, '', '', Entered, '',
         Comments])

    # Save modified workbook
    wb.save(path)


def saveClinicGI(patient, MRNumber, Date, DayType, Const, Dia, Vom, Nausea, Gag, Nissen, ConstDes, DiaDes, VomDes,
                 NauseaDes, GagDes, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Clinic_GI_Issues_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb['Sheet1']
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Create appropriate columns
        ws.append(["MRNumber", "Date", "Day_Type", "Clinic_Const", "Clinic_Dia", "Clinic_Vom", "Clinic_Nausea",
                   "Clinic_Gag_Retch", "Clinic_Nissen", "Clinic_Descr_Const", "Clinic_Descr_Dia", "Clinic_Descr_Vom",
                   "Clinic_Descr_Nausea", "Clinic_Descr_Gag_Retch", "Entered", "Audited", "Comments"])

    # Add data
    ws.append(
        [MRNumber, Date, DayType, Const, Dia, Vom, Nausea, Gag, Nissen, ConstDes, DiaDes, VomDes, NauseaDes, GagDes,
         Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveClinicalLabs(patient, MRNumber, Date, Time, DayType, Source, Fasting, TGBlood, HDLBlood, LDLBlood, TCBlood,
                     NABlood, KBlood, ChlBlood, CO2Blood, BUNBlood, CrBlood, GlusBlood, CaBlood, MagBlood, PhosBlood,
                     UricAcidBlood, ProBlood,
                     AlbBlood, TBilBlood, AlpBlood, AstBlood, AltBlood, RBCBlood, HgbBlood, HctBlood, PlateletBlood,
                     MCVBlood, MCHBlood, MCHCBlood, MPVBlood, RDWBlood, WBCBlood, AmmoniaBlood, BHbBlood, AcacBlood,
                     NeutrophilsBlood, LymphocytesBlood, MonocytesBlood, EosinophilsBlood, BasophilsBlood,
                     LargeUnstainedCellsBlood, NeutrophilsAbsoluteBlood, LymphocytesAbsoluteBlood,
                     MonocytesAbsoluteBlood, EosinophilsAbsoluteBlood, BasophilsAbsoluteBlood,
                     GlusBloodCRC, LactBloodCRCMmol, LabBlood1, LabBlood2, LabBlood3, LabBlood4, LabBlood5, LabBlood6,
                     LabBlood7, LabBlood8, LabBlood9, LabBlood10, LabBlood11, LabBlood12, LabBlood13, LabBlood14,
                     LabBlood15, LabBlood16, LabBlood17,
                     LabBlood18, LabBlood19, LabBlood20, LabBlood21, LabBlood22, LabBlood23, LabBlood24, LabBlood25,
                     LabBlood26, LabBlood27, LabBlood28, LabBlood29, LabBlood30, LabBlood31, LabBlood32, LabBlood33,
                     LabBlood34, LabBlood35, LabBlood36,
                     LabBlood37, LabBlood38, LabBlood39, LabBlood40, LabBlood41, LabBlood42, LabBlood43, LabBlood44,
                     LabBlood45, LabBlood46, LabBlood47, LabBlood48, LabBlood49, LabBlood50, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Clinical_Labs_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Sheet1"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Create appropriate columns
        ws.append(["MRNumber", "Date", "Time", "Day_Type", "Source", "Fasting", "Tg_Blood", "HDL_Blood", "LDL_Blood",
                   "TC_Blood", "Na_Blood", "K_Blood", "Chl_Blood", "CO2_Blood", "BUN_Blood", "Cr_Blood", "Glus_Blood",
                   "Ca_Blood", "Mag_Blood", "Phos_Blood", "Uric_Acid_Blood", "Pro_Blood", "Alb_Blood", "TBil_Blood",
                   "Alp_Blood", "Ast_Blood", "Alt_Blood", "RBC_Blood", "Hgb_Blood", "Hct_Blood", "Platelet_Blood",
                   "MCV_Blood",
                   "MCH_Blood", "MCHC_Blood", "MPV_Blood", "RDW_Blood", "WBC_Blood", "Ammonia_Blood", "BHb_Blood_Mmol",
                   "Acac_Blood", "Neutrophils_Blood", "Lymphocytes_Blood", "Monocytes_Blood", "Eosinophils_Blood",
                   "Basophils_Blood",
                   "Large_Unstained_Cells_Blood", "Neutrophils_Absolute_Blood", "Lymphocytes_Absolute_Blood",
                   "Monocytes_Absolute_Blood", "Eosinophils_Absolute_Blood", "Basophils_Absolute_Blood",
                   "Glus_Blood_CRC", "Lact_Blood_CRC_Mmol",
                   "Lab_Blood_1", "Lab_Blood_2", "Lab_Blood_3", "Lab_Blood_4", "Lab_Blood_5", "Lab_Blood_6",
                   "Lab_Blood_7", "Lab_Blood_8", "Lab_Blood_9", "Lab_Blood_10", "Lab_Blood_11", "Lab_Blood_12",
                   "Lab_Blood_13", "Lab_Blood_14",
                   "Lab_Blood_15", "Lab_Blood_16", "Lab_Blood_17", "Lab_Blood_18", "Lab_Blood_19", "Lab_Blood_20",
                   "Lab_Blood_21", "Lab_Blood_22", "Lab_Blood_23", "Lab_Blood_24", "Lab_Blood_25", "Lab_Blood_26",
                   "Lab_Blood_27",
                   "Lab_Blood_28", "Lab_Blood_29", "Lab_Blood_30", "Lab_Blood_31", "Lab_Blood_32", "Lab_Blood_33",
                   "Lab_Blood_34", "Lab_Blood_35", "Lab_Blood_36", "Lab_Blood_37", "Lab_Blood_38", "Lab_Blood_39",
                   "Lab_Blood_40",
                   "Lab_Blood_41", "Lab_Blood_42", "Lab_Blood_43", "Lab_Blood_44", "Lab_Blood_45", "Lab_Blood_46",
                   "Lab_Blood_47", "Lab_Blood_48", "Lab_Blood_49", "Lab_Blood_50", "Entered", "Audited", "Comments"])

    # Add data
    ws.append([MRNumber, Date, Time, DayType, Source, Fasting, TGBlood, HDLBlood, LDLBlood, TCBlood, NABlood, KBlood,
               ChlBlood, CO2Blood, BUNBlood, CrBlood, GlusBlood, CaBlood, MagBlood, PhosBlood, UricAcidBlood, ProBlood,
               AlbBlood, TBilBlood, AlpBlood, AstBlood, AltBlood, RBCBlood, HgbBlood, HctBlood, PlateletBlood, MCVBlood,
               MCHBlood, MCHCBlood, MPVBlood, RDWBlood, WBCBlood, AmmoniaBlood, BHbBlood, AcacBlood,
               NeutrophilsBlood, LymphocytesBlood, MonocytesBlood, EosinophilsBlood, BasophilsBlood,
               LargeUnstainedCellsBlood, NeutrophilsAbsoluteBlood, LymphocytesAbsoluteBlood, MonocytesAbsoluteBlood,
               EosinophilsAbsoluteBlood, BasophilsAbsoluteBlood,
               GlusBloodCRC, LactBloodCRCMmol, LabBlood1, LabBlood2, LabBlood3, LabBlood4, LabBlood5, LabBlood6,
               LabBlood7, LabBlood8, LabBlood9, LabBlood10, LabBlood11, LabBlood12, LabBlood13, LabBlood14, LabBlood15,
               LabBlood16, LabBlood17,
               LabBlood18, LabBlood19, LabBlood20, LabBlood21, LabBlood22, LabBlood23, LabBlood24, LabBlood25,
               LabBlood26, LabBlood27, LabBlood28, LabBlood29, LabBlood30, LabBlood31, LabBlood32, LabBlood33,
               LabBlood34, LabBlood35, LabBlood36,
               LabBlood37, LabBlood38, LabBlood39, LabBlood40, LabBlood41, LabBlood42, LabBlood43, LabBlood44,
               LabBlood45, LabBlood46, LabBlood47, LabBlood48, LabBlood49, LabBlood50, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveDailyIntake(patient, MRNumber, Date, DayType, PKTNUM, DataQuality, DayQuality, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Daily_Intake_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Daily_Intake"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily_Intake"

        # Create appropriate columns
        ws.append(["MRNumber", "Date", "Day_Type", "PKT_Recipe_Number", "Data_Quality_D", "Day_Quality_D", "Entered",
                   "Audited", "Comments"])

    # Add data
    ws.append([MRNumber, Date, DayType, PKTNUM, DataQuality, DayQuality, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveDietRX(patient, MRNumber, Date, DayType, ROF, RFCD, SnackCal, SnackRatio, SnackNumber, MealNumber, MealRatio,
               TotalCal, Protein, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Diet_RX_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Diet_Rx"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Diet_Rx"

        # Create appropriate columns
        ws.append(["MRNumber", "Date", "Day_Type", "ROF_Pr", "Reason_For_Change_Diet", "Snack_Cal_Pr", "Snack_Ratio_Pr",
                   "Snack_Number_Pr", "Meal_Number_Pr", "Ratio_Pr", "Cal_Pr", "Pro_Pr", "Entered", "Audited",
                   "Comments"])

    # Add data
    ws.append([MRNumber, Date, DayType, ROF, RFCD, SnackCal, SnackRatio, SnackNumber, MealNumber, MealRatio, TotalCal,
               Protein, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveMedData(patient, MRNumber, Date, DayType, NDID, MedID, RFCM, ProdName, DailyDose, MedDose, MedComments, Entered,
                Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Med_Data_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Sheet1"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Create appropriate columns
        ws.append(["MRNumber", "Date", "Day_Type", "NDID", "Med_ID", "Reason_For_Change_Med", "Prod_Name",
                   "Daily_Med_Dose_Mg", "Med_Doses", "Med_Comments", "Entered", "Audited", "Comments"])

    # Add data
    ws.append(
        [MRNumber, Date, DayType, NDID, MedID, RFCM, ProdName, DailyDose, MedDose, MedComments, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveMenus(patient, MRNumber, Date, CalPrcnt, ProcntPrcnt, RatioPr, MealNumber, SnackNumber, RecipeName,
              RecipeNumber, RecipeIngredientAmount, NDID, ProdName, RecipeType, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Menus_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Menus"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Menus"

        # Create appropriate columns
        ws.append(["MRNumber", "Date", "Cal_Prcnt", "Procnt_Prcnt", "Ratio_Pr", "Meal_Number_Pr", "Snack_Number_Pr",
                   "PKT_Recipe_Name", "PKT_Recipe_Number", "PKT_Recipe_Ingredient_Amount", "NDID", "Prod_Name",
                   "Recipe_Type", "Entered", "Audited", "Comments"])

    # Add data
    ws.append([MRNumber, Date, CalPrcnt, ProcntPrcnt, RatioPr, MealNumber, SnackNumber, RecipeName, RecipeNumber,
               RecipeIngredientAmount, NDID, ProdName, RecipeType, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveOtherMed(patient, MRNumber, Date, NDID, ProdName, MedAmount, MedUnit, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Other_Med_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Menus"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Menus"

        # Create appropriate columns
        ws.append(["MRNumber", "Date", "NDID", "Prod_Name", "Med_Amount", "Med_Unit", "Entered", "Audited", "Comments"])

    # Add data
    ws.append([MRNumber, Date, NDID, ProdName, MedAmount, MedUnit, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveSeizureData(patient, MRNumber, Date, DayType, DataQuality, SeizureSeverity, SeizureLength, SeizureType,
                    SeizureVariable, SeizureNumber, SeizureCluster, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Seizure_Data_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Sheet1"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Create appropriate columns
        ws.append(
            ["MRNumber", "Date", "Day_Type", "Data_Quality_S", "Seizure_Severity", "Seizure_Length", "Seizure_Type",
             "Seizure_Variable", "Seizure_Number", "Seizure_Cluster", "Entered", "Audited", "Comments"])

    # Add data
    ws.append([MRNumber, Date, DayType, DataQuality, SeizureSeverity, SeizureLength, SeizureType, SeizureVariable,
               SeizureNumber, SeizureCluster, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveSeizureRanking(patient, MRNumber, SeizureParameter, SeizureEntry, SeizureRanking, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Seizure_Ranking_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Sheet1"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Create appropriate columns
        ws.append(
            ["MRNumber", "Seizure_Parameter", "Seizure_Entry", "Seizure_Ranking", "Entered", "Audited", "Comments"])

    # Add data
    ws.append([MRNumber, SeizureParameter, SeizureEntry, SeizureRanking, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveUrineKtSG(patient, MRNumber, Date, DayType, UrineKt, UrineSG, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Urine_Kt_SG_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Sheet1"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Create appropriate columns
        ws.append(["MRNumber", "Date", "Day_Type", "Urine_Kt", "Urine_SG", "Entered", "Audited", "Comments"])

    # Add data
    ws.append([MRNumber, Date, DayType, UrineKt, UrineSG, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveVitals(patient, MRNumber, Date, DayType, Source, BPSys, BPDia, Temp, RR, HR, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Vitals_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Sheet1"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Create appropriate columns
        ws.append(["MRNumber", "Date", "Day_Type", "Source", "BP_Sys", "BP_Dia", "T", "RR", "HR", "Entered", "Audited",
                   "Comments"])

    # Add data
    ws.append([MRNumber, Date, DayType, Source, BPSys, BPDia, Temp, RR, HR, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)


def saveVNS(patient, MRNumber, Date, DayType, MRMagAct, OutPutCurr, VNSFrequency, PluseWidth, SignalOn, SingnalOff,
            MagnetCurrent, MagnetOn, MagnetPulse, LeadTest, Entered, Comments):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Clinic_VNS_Source.xlsx"

    # Tries to load workbook
    try:
        wb = load_workbook(path)
        ws = wb["Sheet1"]
    # On failure ensures they have the right directories using setNewPatient
    except FileNotFoundError:
        setNewPatient(patient)

        # Create new workbook and name sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Create appropriate columns
        ws.append(
            ["MRNumber", "Date", "Mag_Act", "Output_Curr", "VNS_Frequency", "Pulse_Width", "Signal_On", "Signal_Off",
             "Magnet_Current", "Magnet_On", "Magnet_Pulse", "Lead_Test" "Entered", "Audited", "Comments"])

    # Add data
    ws.append([patient, MRNumber, Date, DayType, MRMagAct, OutPutCurr, VNSFrequency, PluseWidth, SignalOn, SingnalOff,
               MagnetCurrent, MagnetOn, MagnetPulse, LeadTest, Entered, "", Comments])

    # Save modified workbook
    wb.save(path)

################### Getter Methods for excel sheets ###################

# def getDemographics(mrn):
#    print("in demographics")
#    try:
#        path = str(os.getcwd()) + r"\Current Patients\Demographics_Source.xlsx"
#        workbook = xlrd.open_workbook(filename=path)
#        sheetnames = workbook.sheet_names()
#        worksheet = workbook.sheet_by_name(sheetnames[0])
#        for row_ndx in range(1, worksheet.nrows):
#            for col_ndx in range(0, 1):
#                cell_val = int(worksheet.cell_value(row_ndx, col_ndx))
#                print ('Column: [%s] cell_obj: [%s]' % (col_ndx, cell_val))
#        return 1
#    except:
#        exit("demo error")

# Returns dataframe containing data inside the excel sheet
def getAnthropometricsDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Anthropometrics_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe


def getAnthropometricsClinicalDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Anthropometrics_Clinical.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getVNSDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Clinic_VNS_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getVitalsDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Vitals_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getUrineKtSgDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Urine_Kt_SG_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getSeizureRankingDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Seizure_Ranking_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getSeizureDataDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Seizure_Data_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getOtherMedDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Other_Med_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getMenusDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Menus_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getMedDataDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Med_Data_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getDietRXDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Diet_RX_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getDailyIntakeDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Daily_Intake_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getClinicalLabsDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Clinical_Labs_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getClinicGIDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Clinic_GI_Issues_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe

def getAlertnessDataFrame(patient):
    path = r"Current Patients\\" + patient + r"\\DataBases\\Data\\" + patient + r"_Alertness_Source.xlsx"
    dataframe = pandas.read_excel(path)
    return dataframe
    